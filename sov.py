#!/usr/bin/env python

import re
import json
import argparse

from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser(
        description = "Grabs vote totals from the statement of the vote xlsx file")
    parser.add_argument("input_filename")
    parser.add_argument("-o", "--output-json-filename",
                        default="data/precinct_vote_totals.json")
    args = parser.parse_args()

    wb = load_workbook(filename=args.input_filename, read_only=True)
    ws = wb.worksheets[0]

    i = 5
    pct = None
    pct_data = {}
    while True:
        val = ws.cell(i, 1).value
        if not val:
            break
        elif val.startswith('PCT'):
            pct = re.match(r"PCT (\d+).*", val).group(1)
        elif val == "Total":
            pct_data[pct] = int(ws.cell(i, 4).value)
            print(pct, pct_data[pct])

        i += 1

    with open(args.output_json_filename, "w") as f:
        json.dump(pct_data, f)

if __name__ == "__main__":
    main()
